# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


import requests
from bs4 import BeautifulSoup
import re
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import info
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import os
import time
import json
from datetime import datetime


root_url = 'https://partner.caravanpark.kr'
ID = 'elpark4'
PW = '1234'

def setup_selenium_and_login(user_id, user_password):
    """
    Selenium WebDriver를 설정하고, 로그인 성공 페이지 URL('/reservation/monthly')로
    이동하는 것을 기준으로 안정적으로 로그인을 확인합니다.
    """
    print("[DEBUG] Setting up Selenium WebDriver and logging in...")
    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    # options.add_argument('--headless') # 필요 시 주석 해제

    driver = None
    try:
        # 1. webdriver-manager 사용 시도
        service = ChromeService(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
    except Exception as e:
        print(f"[DEBUG] webdriver-manager failed: {e}")
        # 2. 로컬 chromedriver 사용 시도
        chromedriver_path = "chromedriver-win64/chromedriver.exe"
        if os.path.exists(chromedriver_path):
            print(f"[DEBUG] Using local chromedriver: {chromedriver_path}")
            service = ChromeService(executable_path=chromedriver_path)
            driver = webdriver.Chrome(service=service, options=options)
        else:
            raise Exception("ChromeDriver not found by webdriver-manager or in local path.")

    # 로그인 페이지로 이동 및 로그인 수행
    signin_page_url = f"{root_url}/auth/signin"
    driver.get(signin_page_url)
    
    try:
        # 필드가 나타날 때까지 최대 10초 대기
        username_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "masterId"))
        )
        password_field = driver.find_element(By.NAME, "masterPass")
        
        username_field.send_keys(user_id)
        password_field.send_keys(user_password)

        login_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
        login_button.click()
        
        print("[DEBUG] 로그인 버튼 클릭. '/reservation/monthly' 페이지로 이동할 때까지 최대 15초 대기...")
        
        # 가장 확실한 방법: 로그인 성공 후의 실제 URL로 이동했는지 확인
        WebDriverWait(driver, 15).until(
            EC.url_contains('/reservation/monthly')
        )
        
        print(f"[DEBUG] 로그인 성공! 성공 페이지로 이동했습니다: {driver.current_url}")

    except TimeoutException:
        print("[CRITICAL ERROR] 로그인 실패: 제한 시간(15초) 내에 예약 관리 페이지로 이동하지 못했습니다.")
        print(f"[DEBUG] 현재 머물러 있는 페이지 주소: {driver.current_url}")
        print("[INFO] 아이디/패스워드가 정확한지, 또는 사이트의 다른 문제가 있는지 확인해주세요.")
        if driver:
            driver.quit()
        raise Exception("로그인에 실패했습니다. 사이트 응답이 없거나 인증 정보가 틀렸을 수 있습니다.")
    except Exception as e:
        if driver:
            driver.quit()
        raise Exception(f"로그인 중 예상치 못한 오류 발생: {e}")

    return driver

def get_reservation():
    """
    사용자께서 제공해주신 HTML 구조를 기반으로, 가장 확실한 대기 방법과
    정확한 파싱 로직으로 모든 예약 정보를 가져옵니다.
    """
    today_str = datetime.now().strftime('%Y-%m-%d')
    driver = None
    all_reservations = []

    try:
        driver = setup_selenium_and_login(ID, PW)

        print("\n[INFO] 예약 목록 페이지로 이동하여 검색을 시작합니다...")
        list_url = f"{root_url}/reservation/list?startDate={today_str}&endDate={today_str}&dateMode=sdate"
        driver.get(list_url)

        # 체크박스들을 제어 (가장 강력하고 안정적인 방식으로 수정)
        checkbox_labels_to_check = ["예약완료", "입금대기"]
        for label_text in checkbox_labels_to_check:
            try:
                label = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, f"//label[contains(normalize-space(), '{label_text}')]"))
                )
                checkbox = label.find_element(By.XPATH, ".//input[@type='checkbox']")
                
                if not checkbox.is_selected():
                    driver.execute_script("arguments[0].click();", checkbox)
                    WebDriverWait(driver, 2).until(EC.element_to_be_selected(checkbox))
                    print(f"[DEBUG] '{label_text}' 체크박스를 클릭했습니다.")
                else:
                    print(f"[DEBUG] '{label_text}' 체크박스가 이미 선택되어 있습니다.")
            except Exception as e:
                print(f"[WARN] '{label_text}' 체크박스 처리 중 오류 발생: {e}")
                # 오류가 발생해도 계속 진행

        # '검색' 버튼 클릭 (가장 강력한 방식으로 수정)
        try:
            print("[DEBUG] '검색' 버튼을 클릭합니다.")
            search_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(., '검색')]"))
            )
            driver.execute_script("arguments[0].click();", search_button)
        except Exception as e:
            print("[CRITICAL ERROR] '검색' 버튼을 클릭하는 데 실패했습니다.")
            raise e

        # --- 데이터 파싱 로직 시작 ---
        try:
            # 최종 해결책: 첫 번째 예약 번호 링크가 나타날 때까지 대기
            print("[DEBUG] 검색 결과를 기다립니다 (테이블의 첫 예약번호 링크가 보일 때까지 최대 15초)...")
            WebDriverWait(driver, 15).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "tbody a[href*='/detail/']"))
            )
            print("[DEBUG] 예약 목록 테이블이 표시된 것을 확인했습니다.")
        except TimeoutException:
            print("[INFO] 검색 결과, 오늘 날짜의 예약이 없습니다.")
            if driver:
                driver.quit()
            return [], today_str

        # 페이지 소스를 가져와서 BeautifulSoup으로 파싱
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')

        # 사용자께서 알려주신 정확한 선택자 사용
        table_body = soup.find('tbody', class_='divide-y')

        if not table_body:
            print("[INFO] 예약 테이블(tbody)을 찾았으나 내용이 없습니다.")
            if driver:
                driver.quit()
            return [], today_str

        # 모든 행(tr)을 가져옴
        rows = table_body.find_all('tr', recursive=False)
        print(f"[DEBUG] 테이블에서 {len(rows)}개의 행(tr)을 발견했습니다.")

        for row in rows:
            # 상세 정보 행(colspan이 있는)은 건너뜀
            if row.find('td', colspan=True):
                continue
            
            cols = row.find_all('td')
            if len(cols) < 16:
                print(f"[DEBUG] 데이터 행이 아닌 것 같아 건너뜁니다.")
                continue

            try:
                # 제공해주신 HTML 구조에 기반하여 정확한 인덱스로 데이터 추출
                reservation_data = {
                    '예약번호': cols[0].get_text(strip=True),
                    '예약자명': cols[1].get_text(separator=' ', strip=True),
                    '객실명': cols[2].get_text(strip=True),
                    '숙박일': cols[3].get_text(separator=' ', strip=True),
                    '기간': cols[4].get_text(strip=True),
                    '인원': cols[5].get_text(separator=' ', strip=True),
                    '판매금액': cols[10].get_text(strip=True),
                    '상태': cols[14].get_text(strip=True),
                    '예약경로': cols[15].get_text(strip=True),
                }
                all_reservations.append(reservation_data)
            except (AttributeError, IndexError) as e:
                print(f"[WARN] 데이터 행 파싱 중 오류 발생 (예상치 못한 구조): {e}")
                continue
        
        print(f"[INFO] 총 {len(all_reservations)}건의 예약 정보를 성공적으로 파싱했습니다.")
        return all_reservations, today_str

    except Exception as e:
        print(f"[CRITICAL ERROR] 스크립트 실행 중 심각한 오류가 발생했습니다: {e}")
        # 디버깅을 위해 스크린샷 저장
        if driver:
            error_screenshot = f"error_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            driver.save_screenshot(error_screenshot)
            print(f"[DEBUG] 오류 발생 당시 화면을 '{error_screenshot}' 파일로 저장했습니다.")
    finally:
        if driver:
            driver.quit()
            print("[INFO] WebDriver를 종료했습니다.")
            
    return all_reservations, today_str


def make_daily_paper():
    reservations, today_str = get_reservation()

    if not reservations:
        print("처리할 예약 데이터가 없습니다.")
        return

    df = pd.DataFrame(reservations)

    filename = f"예약현황_{today_str}.xlsx"
    # Using openpyxl to create and style the workbook
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "예약 현황"

        # DataFrame to Excel
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Get the dimensions of the sheet
        max_row = ws.max_row
        max_col = ws.max_column

        # --- Styling ---
        # 1. Font
        header_font = Font(name='맑은 고딕', size=11, bold=True)
        body_font = Font(name='맑은 고딕', size=10)

        # 2. Alignment
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 3. Border
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Apply styles
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border
                if cell.row == 1:
                    cell.font = header_font
                else:
                    cell.font = body_font
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Save the file
        wb.save(filename)
        print(f"\n[SUCCESS] '{filename}' 파일이 성공적으로 생성되었습니다.")
        print(f"총 {len(reservations)}개의 예약 정보를 처리했습니다.")

    except Exception as e:
        print(f"[CRITICAL ERROR] 엑셀 파일 생성 중 오류가 발생했습니다: {e}")
        # 오류 발생 시, 간단한 CSV 파일로 저장 시도
        try:
            csv_filename = f"예약현황_백업_{today_str}.csv"
            df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
            print(f"[INFO] 대신 '{csv_filename}' CSV 파일로 데이터를 저장했습니다.")
        except Exception as csv_e:
            print(f"[CRITICAL ERROR] CSV 백업 파일 저장에도 실패했습니다: {csv_e}")


# def today_reservation(room_type, client_name, phone_number):

def main():
    """메인 실행 함수"""
    try:
        make_daily_paper()
    except Exception as e:
        print(f"\n[CRITICAL ERROR] 스크립트 실행 중 예외가 발생했습니다: {e}")
    finally:
        print("\n\n✅ 작업 완료! 이제 터미널을 닫으셔도 됩니다.\n")

if __name__ == "__main__":
    main()
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
